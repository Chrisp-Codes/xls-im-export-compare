import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill
import os
import datetime

# ---------------------------------------------
# 1) Hilfsfunktionen
# ---------------------------------------------

def normalize_name(name_string: str) -> str:
    """
    Normalisiert einen String für den Vergleich von "Namensfeldern".
    - Kleinschreibung
    - Entfernt Leerzeichen und Bindestriche
    """
    if not isinstance(name_string, str):
        return ""
    result = name_string.lower()
    for char in [" ", "-"]:
        result = result.replace(char, "")
    return result

def color_cell(ws, row_idx, col_idx, color="FF0000"):
    """Färbt eine einzelne Zelle in der angegebenen Farbe (Standard: Rot)."""
    cell = ws.cell(row=row_idx, column=col_idx)
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.fill = fill

def color_row(ws, row_idx, max_cols, color="FF0000"):
    """Färbt eine ganze Zeile (1 bis max_cols) in der angegebenen Farbe (Standard: Rot)."""
    for col_idx in range(1, max_cols + 1):
        color_cell(ws, row_idx, col_idx, color=color)

def to_str_with_date_check(value):
    """
    Wandelt ein Value in einen einheitlichen String um.
    - Falls datetime.date/datetime.datetime: Format "TT.MM.YYYY"
    - Sonst .strip()
    """
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%d.%m.%Y")
    return str(value).strip() if value else ""

def match_employee(import_row, export_dict, import_headers, export_headers):
    """
    Versucht anhand (Personalnummer, Vorname, Nachname) einen passenden Eintrag
    im export_dict zu finden. (mind. 2 der 3 Merkmale müssen passen).
    
    - import_row: Liste der Zellwerte in der Import-Zeile
    - export_dict: dict mit Schlüsseln (persnr_norm, vorname_norm, nachname_norm)
    - import_headers, export_headers: "Spaltenname -> Spaltenindex (1-basiert)"
    """
    persnr_val = ""
    vorname_val = ""
    nachname_val = ""
    
    if "Personalnummer" in import_headers:
        persnr_val = import_row[ import_headers["Personalnummer"] - 1 ]
    if "Vorname*" in import_headers:
        vorname_val = import_row[ import_headers["Vorname*"] - 1 ]
    if "Nachname" in import_headers:
        nachname_val = import_row[ import_headers["Nachname"] - 1 ]
    
    persnr_str = str(persnr_val).strip().lower() if persnr_val else ""
    vorname_norm = normalize_name(vorname_val)
    nachname_norm = normalize_name(nachname_val)
    
    matches = []
    for (exp_persnr, exp_vorname, exp_nachname), export_rows in export_dict.items():
        match_count = 0
        
        if persnr_str and persnr_str == exp_persnr:
            match_count += 1
        if vorname_norm and vorname_norm == exp_vorname:
            match_count += 1
        if nachname_norm and nachname_norm == exp_nachname:
            match_count += 1
        
        if match_count >= 2:
            for row_data in export_rows:
                matches.append(row_data)
    
    return matches

# ---------------------------------------------
# 2) Vergleichs-Funktion
# ---------------------------------------------

def compare_excels(import_file, export_file, output_file):
    """
    Liest zwei Excel-Dateien ein, vergleicht sie und markiert in der Import-Datei
    die Abweichungen (rote Zellen/Zeilen) bzw. Mehrfachtreffer (orange Zeilen).
    Anschließend wird das Ergebnis unter output_file gespeichert.
    """
    import_wb = openpyxl.load_workbook(import_file)
    export_wb = openpyxl.load_workbook(export_file)
    
    import_ws = import_wb.active
    export_ws = export_wb.active
    
    # Kopfzeilen einlesen
    import_headers = {}
    export_headers = {}
    
    max_col_import = import_ws.max_column
    max_row_import = import_ws.max_row
    max_col_export = export_ws.max_column
    max_row_export = export_ws.max_row
    
    for col_idx in range(1, max_col_import + 1):
        val = import_ws.cell(row=1, column=col_idx).value
        if val:
            import_headers[str(val).strip()] = col_idx
    
    for col_idx in range(1, max_col_export + 1):
        val = export_ws.cell(row=1, column=col_idx).value
        if val:
            export_headers[str(val).strip()] = col_idx
    
    # Schnittmenge der Spalten
    common_columns = set(import_headers.keys()).intersection(set(export_headers.keys()))
    
    # 2.1) Bestimmte Spalten = "Namenslogik"
    NAME_COLUMNS = [
        "Vorname*",
        "Vorsatzwort",
        "Nachname",
        "Namenszusatz",
        "Titel",
        "Krankenkasse"
    ]
    
    # 2.2) Datumsspalten
    DATE_COLUMNS = [
        "Geburtstag",
        "Geburtsdatum",
        # Falls du weitere Datumsspalten hast, hier ergänzen
    ]
    
    # 2.3) Bestimmte Spalten = "case-insensitive" (z.B. Familienstand, Elterneigenschaft)
    CASE_INSENSITIVE_COLUMNS = [
        "Familienstand",
        "Elterneigenschaft"
    ]
    
    # Export-Daten in Dictionary (für Schnellsuche)
    persnr_col = export_headers.get("Personalnummer", None)
    vorname_col = export_headers.get("Vorname*", None)
    nachname_col = export_headers.get("Nachname", None)
    
    export_dict = {}
    for row_idx in range(2, max_row_export + 1):
        row_values = [
            export_ws.cell(row=row_idx, column=col_idx).value
            for col_idx in range(1, max_col_export + 1)
        ]
        
        if persnr_col:
            persnr_val = export_ws.cell(row=row_idx, column=persnr_col).value
            persnr_norm = str(persnr_val).strip().lower() if persnr_val else ""
        else:
            persnr_norm = ""
        
        if vorname_col:
            vorname_val = export_ws.cell(row=row_idx, column=vorname_col).value
            vorname_norm = normalize_name(vorname_val)
        else:
            vorname_norm = ""
        
        if nachname_col:
            nachname_val = export_ws.cell(row=row_idx, column=nachname_col).value
            nachname_norm = normalize_name(nachname_val)
        else:
            nachname_norm = ""
        
        key = (persnr_norm, vorname_norm, nachname_norm)
        if key not in export_dict:
            export_dict[key] = []
        export_dict[key].append((row_idx, row_values))
    
    # Import-Daten zeilenweise prüfen
    for row_idx in range(2, max_row_import + 1):
        row_values = [
            import_ws.cell(row=row_idx, column=col_idx).value
            for col_idx in range(1, max_col_import + 1)
        ]
        
        matches = match_employee(row_values, export_dict, import_headers, export_headers)
        
        if len(matches) == 0:
            # Kein passender Datensatz -> ganze Zeile rot
            color_row(import_ws, row_idx, max_col_import, color="FF0000")
        
        elif len(matches) > 1:
            # Mehrfachtreffer -> ganze Zeile orange
            color_row(import_ws, row_idx, max_col_import, color="FFA500")
        
        else:
            # Genau ein Treffer
            export_row_idx, export_row_vals = matches[0]
            
            # Jetzt spaltenweise vergleichen
            for col_name in common_columns:
                import_col_idx = import_headers[col_name]
                export_col_idx = export_headers[col_name]
                
                import_val = import_ws.cell(row=row_idx, column=import_col_idx).value
                export_val = export_ws.cell(row=export_row_idx, column=export_col_idx).value
                
                # 1) Namens-Logik (Groß/Klein + Leer-/Bindestriche ignorieren)
                if col_name in NAME_COLUMNS:
                    if normalize_name(import_val) != normalize_name(export_val):
                        color_cell(import_ws, row_idx, import_col_idx, color="FF0000")
                
                # 2) E-Mail: nur Groß-/Kleinschreibung ignorieren
                elif col_name == "E-Mail":
                    i_val = str(import_val).strip().lower() if import_val else ""
                    e_val = str(export_val).strip().lower() if export_val else ""
                    if i_val != e_val:
                        color_cell(import_ws, row_idx, import_col_idx, color="FF0000")
                
                # 3) Datumsfelder: in String "TT.MM.YYYY" umwandeln und dann vergleichen
                elif col_name in DATE_COLUMNS:
                    i_val = to_str_with_date_check(import_val)
                    e_val = to_str_with_date_check(export_val)
                    if i_val != e_val:
                        color_cell(import_ws, row_idx, import_col_idx, color="FF0000")
                
                # 4) Case-Insensitive Spalten (z.B. Familienstand, Elterneigenschaft)
                elif col_name in CASE_INSENSITIVE_COLUMNS:
                    i_val = str(import_val).strip().lower() if import_val else ""
                    e_val = str(export_val).strip().lower() if export_val else ""
                    if i_val != e_val:
                        color_cell(import_ws, row_idx, import_col_idx, color="FF0000")
                
                # 5) Standard-Vergleich (exakt, nur Leerzeichen trimmen)
                else:
                    i_val = str(import_val).strip() if import_val else ""
                    e_val = str(export_val).strip() if export_val else ""
                    if i_val != e_val:
                        color_cell(import_ws, row_idx, import_col_idx, color="FF0000")
    
    # Speichern
    import_wb.save(output_file)

# ---------------------------------------------
# 3) GUI-Anwendung mit tkinter
# ---------------------------------------------

class ExcelCompareApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel Compare Tool")
        
        # Variablen für Pfade
        self.import_path = tk.StringVar()
        self.export_path = tk.StringVar()
        self.output_path = tk.StringVar()
        
        # Labels und Eingabefelder + Buttons
        tk.Label(master, text="Import-Datei:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        tk.Entry(master, textvariable=self.import_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(master, text="Durchsuchen...", command=self.browse_import).grid(row=0, column=2, padx=5, pady=5)
        
        tk.Label(master, text="Export-Datei:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        tk.Entry(master, textvariable=self.export_path, width=50).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(master, text="Durchsuchen...", command=self.browse_export).grid(row=1, column=2, padx=5, pady=5)
        
        tk.Label(master, text="Ergebnis-Datei:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        tk.Entry(master, textvariable=self.output_path, width=50).grid(row=2, column=1, padx=5, pady=5)
        tk.Button(master, text="Speichern unter...", command=self.browse_output).grid(row=2, column=2, padx=5, pady=5)
        
        # Start-Button
        tk.Button(master, text="Vergleich starten",
                  command=self.start_compare, fg="white", bg="green").grid(row=3, column=0, columnspan=3, pady=10)
    
    def browse_import(self):
        path = filedialog.askopenfilename(
            title="Import-Datei auswählen",
            filetypes=[("Excel-Dateien", "*.xlsx *.xlsm *.xls")]
        )
        if path:
            self.import_path.set(path)
    
    def browse_export(self):
        path = filedialog.askopenfilename(
            title="Export-Datei auswählen",
            filetypes=[("Excel-Dateien", "*.xlsx *.xlsm *.xls")]
        )
        if path:
            self.export_path.set(path)
    
    def browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Ergebnis-Datei speichern",
            defaultextension=".xlsx",
            filetypes=[("Excel-Datei", "*.xlsx")]
        )
        if path:
            self.output_path.set(path)
    
    def start_compare(self):
        imp = self.import_path.get()
        exp = self.export_path.get()
        out = self.output_path.get()
        
        if not os.path.isfile(imp):
            messagebox.showerror("Fehler", "Ungültige Import-Datei.")
            return
        if not os.path.isfile(exp):
            messagebox.showerror("Fehler", "Ungültige Export-Datei.")
            return
        if not out:
            messagebox.showerror("Fehler", "Bitte einen Ausgabepfad festlegen.")
            return
        
        # Vergleich durchführen
        try:
            compare_excels(imp, exp, out)
            messagebox.showinfo("Fertig", "Vergleich abgeschlossen!\nBitte prüfen Sie die Ergebnisdatei.")
        except Exception as e:
            messagebox.showerror("Fehler beim Vergleich", str(e))

def main():
    root = tk.Tk()
    app = ExcelCompareApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
